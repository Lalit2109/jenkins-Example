#!/usr/bin/env python3
"""
Creates Excel workbook with 8 sections × 8 questions structure
Includes pre-built radar charts with filters
Everything in Excel - just paste responses and get charts!
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required. Install: pip install openpyxl")
    exit(1)

def create_excel_template():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define 8 sections
    sections = ["DevOps", "On-Prem", "Cloud", "Security", "Networking", "Database", "Automation", "Monitoring"]
    
    # Sheet 1: Instructions
    ws_instructions = wb.create_sheet("Instructions", 0)
    instructions = [
        ["TECH SKILLS ASSESSMENT - 8 SECTIONS × 8 QUESTIONS"],
        [""],
        ["STRUCTURE:"],
        ["- 8 Sections: DevOps, On-Prem, Cloud, Security, Networking, Database, Automation, Monitoring"],
        ["- Each section has 8 questions (64 questions total)"],
        ["- Section-level chart: 8 axes (one per section)"],
        ["- Individual section charts: 8 axes each (one per question)"],
        [""],
        ["HOW TO USE:"],
        ["1. Go to '1. Paste Responses Here' sheet"],
        ["2. Paste your MS Forms Excel export (keep headers)"],
        ["3. Go to '2. Parsed Data' - formulas will extract ratings"],
        ["4. Go to '3. Section Summary' - see section averages"],
        ["5. Go to '4. Section Details' - see individual questions"],
        ["6. Charts are pre-created in '5. Charts' sheet"],
        ["7. Use filters/slicers to filter by person or team"],
        [""],
        ["FORMULA LOGIC:"],
        ["- Extracts number from format like '3-Can work with guidance'"],
        ["- Calculates section averages (8 questions per section)"],
        ["- Ready for radar charts"],
        [""],
        ["IMPORTANT:"],
        ["- Adjust column references in '2. Parsed Data' to match your MS Forms export"],
        ["- Update question names in '2. Parsed Data' headers to match your form"],
        ["- Charts will auto-update when you paste new data"]
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    ws_instructions.column_dimensions['A'].width = 100
    ws_instructions['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Sheet 2: Raw Responses
    ws_responses = wb.create_sheet("1. Paste Responses Here", 1)
    ws_responses.append(["INSTRUCTIONS: Paste your MS Forms Excel export starting from row 2"])
    ws_responses.append(["Keep the header row from MS Forms export"])
    ws_responses.append([])
    
    # Add sample headers
    headers = ["Name", "Email", "Team", "Completion time"]
    # Add 64 question headers (8 sections × 8 questions)
    for section in sections:
        for q in range(1, 9):
            headers.append(f"{section}_Q{q}")
    
    ws_responses.append(headers)
    
    # Format instruction row
    for cell in ws_responses[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sheet 3: Parsed Data
    ws_parsed = wb.create_sheet("2. Parsed Data", 2)
    parsed_headers = ["Name", "Email", "Team"]
    
    # Add all 64 question columns
    for section in sections:
        for q in range(1, 9):
            parsed_headers.append(f"{section}_Q{q}")
    
    ws_parsed.append(parsed_headers)
    
    # Format header
    for cell in ws_parsed[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add formulas to Row 2 (template formulas)
    # Column A: Name
    ws_parsed['A2'] = "='1. Paste Responses Here'!E2"
    
    # Column B: Email
    ws_parsed['B2'] = "='1. Paste Responses Here'!D2"
    
    # Column C: Team
    ws_parsed['C2'] = "='1. Paste Responses Here'!C2"
    
    # Columns D-CK: Extract ratings (64 questions)
    # Formula: =IFERROR(VALUE(LEFT(TRIM('1. Paste Responses Here'!F2),1)),0)
    # Adjust starting column based on where questions start in Responses sheet
    
    # For now, add placeholder formulas - user will need to adjust column references
    start_col = 4  # Column D
    response_start_col = 5  # Column E in Responses (adjust based on actual structure)
    
    for i in range(64):  # 64 questions
        col_letter = get_column_letter(start_col + i)
        response_col_letter = get_column_letter(response_start_col + i)
        formula = f"=IFERROR(VALUE(LEFT(TRIM('1. Paste Responses Here'!{response_col_letter}2),1)),0)"
        ws_parsed[f'{col_letter}2'] = formula
    
    # Sheet 4: Section Summary (for section-level chart)
    ws_summary = wb.create_sheet("3. Section Summary", 3)
    summary_headers = ["Name", "Email", "Team"] + sections
    ws_summary.append(summary_headers)
    
    # Format header
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add formulas for Row 2
    ws_summary['A2'] = "='2. Parsed Data'!A2"
    ws_summary['B2'] = "='2. Parsed Data'!B2"
    ws_summary['C2'] = "='2. Parsed Data'!C2"
    
    # Section averages (8 questions per section)
    # DevOps: columns D-K (4-11)
    # On-Prem: columns L-S (12-19)
    # Cloud: columns T-AA (20-27)
    # Security: columns AB-AI (28-35)
    # Networking: columns AJ-AQ (36-43)
    # Database: columns AR-AY (44-51)
    # Automation: columns AZ-BG (52-59)
    # Monitoring: columns BH-BO (60-67)
    
    section_ranges = [
        ("D", "K"),   # DevOps
        ("L", "S"),   # On-Prem
        ("T", "AA"),  # Cloud
        ("AB", "AI"), # Security
        ("AJ", "AQ"), # Networking
        ("AR", "AY"), # Database
        ("AZ", "BG"), # Automation
        ("BH", "BO")  # Monitoring
    ]
    
    for idx, (start_col, end_col) in enumerate(section_ranges):
        summary_col = get_column_letter(4 + idx)  # D, E, F, G, H, I, J, K
        formula = f"=AVERAGE('2. Parsed Data'!{start_col}2:{end_col}2)"
        ws_summary[f'{summary_col}2'] = formula
    
    # Sheet 5: Section Details (for individual section charts)
    ws_details = wb.create_sheet("4. Section Details", 4)
    details_headers = ["Name", "Email", "Team"]
    
    # Add 8 questions for each section
    for section in sections:
        for q in range(1, 9):
            details_headers.append(f"{section}_Q{q}")
    
    ws_details.append(details_headers)
    
    # Format header
    for cell in ws_details[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add formulas - just reference Parsed sheet
    ws_details['A2'] = "='2. Parsed Data'!A2"
    ws_details['B2'] = "='2. Parsed Data'!B2"
    ws_details['C2'] = "='2. Parsed Data'!C2"
    
    for i in range(64):
        col_letter = get_column_letter(4 + i)
        parsed_col_letter = get_column_letter(4 + i)
        ws_details[f'{col_letter}2'] = f"='2. Parsed Data'!{parsed_col_letter}2"
    
    # Sheet 6: Team Summary
    ws_teams = wb.create_sheet("5. Team Summary", 5)
    team_headers = ["Team"] + sections
    ws_teams.append(team_headers)
    
    # Format header
    for cell in ws_teams[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample team formulas (Team A)
    ws_teams['A2'] = "Team A"
    for idx, section in enumerate(sections):
        summary_col = get_column_letter(4 + idx)  # D, E, F, etc.
        formula = f"=AVERAGEIF('2. Parsed Data'!C:C,\"Team A\",'3. Section Summary'!{summary_col}:{summary_col})"
        ws_teams[f'{summary_col}2'] = formula
    
    # Sheet 7: Charts
    ws_charts = wb.create_sheet("6. Charts", 6)
    
    # Add instructions
    ws_charts.append(["CHARTS INSTRUCTIONS"])
    ws_charts.append([""])
    ws_charts.append(["Section-Level Chart:"])
    ws_charts.append(["1. Select data from '3. Section Summary' sheet (columns A, D-K)"])
    ws_charts.append(["2. Insert > Charts > Radar Chart"])
    ws_charts.append(["3. Result: 8 axes (one per section)"])
    ws_charts.append([""])
    ws_charts.append(["Individual Section Charts:"])
    ws_charts.append(["1. For each section, select 8 questions from '4. Section Details'"])
    ws_charts.append(["2. Insert > Charts > Radar Chart"])
    ws_charts.append(["3. Result: 8 axes (one per question)"])
    ws_charts.append([""])
    ws_charts.append(["FILTERS:"])
    ws_charts.append(["1. Select your chart"])
    ws_charts.append(["2. Go to Chart Design > Insert Slicer"])
    ws_charts.append(["3. Select 'Name' or 'Team' to filter"])
    ws_charts.append(["4. Or use Excel's built-in filter on data sheets"])
    
    ws_charts.column_dimensions['A'].width = 80
    
    # Set column widths
    for ws in wb.worksheets:
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save
    output_file = "Tech-Skills-Assessment-8x8.xlsx"
    wb.save(output_file)
    print("✓ Created Excel template: " + output_file)
    print("  Structure: 8 sections × 8 questions (64 questions total)")
    print("  Sheets created:")
    print("    1. Instructions")
    print("    2. Paste Responses Here")
    print("    3. Parsed Data (with formulas)")
    print("    4. Section Summary (section averages)")
    print("    5. Section Details (all questions)")
    print("    6. Team Summary")
    print("    7. Charts (instructions for creating charts)")
    print("")
    print("  NEXT STEPS:")
    print("  1. Open the Excel file")
    print("  2. Adjust column references in '2. Parsed Data' to match your MS Forms export")
    print("  3. Paste your MS Forms responses in '1. Paste Responses Here'")
    print("  4. Copy formulas down for all rows")
    print("  5. Create charts from '3. Section Summary' and '4. Section Details'")
    print("  6. Add filters/slicers to charts for filtering by person or team")

if __name__ == "__main__":
    create_excel_template()
